import pytest
import pandas as pd
import io
from datetime import datetime
from unittest.mock import patch
from utils.exporter import (
    get_report_title, export_to_excel, export_html_report,
    generate_html_report, _is_blank_image
)


class TestReportTitle:
    def test_basic_title(self):
        title, subtitle = get_report_title("销售概览", datetime(2025, 1, 1), datetime(2025, 6, 30))
        assert "销售概览" in title
        assert "2025-01-01" in subtitle
        assert "2025-06-30" in subtitle

    def test_title_with_categories(self):
        title, subtitle = get_report_title(
            "品类分析", datetime(2025, 1, 1), datetime(2025, 6, 30),
            categories=["电子产品", "服装"]
        )
        assert "电子产品" in subtitle
        assert "服装" in subtitle

    def test_title_no_filters(self):
        title, subtitle = get_report_title("概览", datetime(2025, 1, 1), datetime(2025, 6, 30))
        assert "时间" in subtitle

    def test_title_with_regions(self):
        title, subtitle = get_report_title(
            "地区", datetime(2025, 1, 1), datetime(2025, 6, 30),
            regions=["北京", "上海"]
        )
        assert "北京" in subtitle

    def test_title_with_channels(self):
        title, subtitle = get_report_title(
            "渠道", datetime(2025, 1, 1), datetime(2025, 6, 30),
            channels=["APP", "小程序"]
        )
        assert "APP" in subtitle


class TestExcelExport:
    def test_export_normal_data(self, sample_orders_df):
        cat_agg = sample_orders_df.groupby('product_category').agg(
            销售额=('amount', 'sum')
        ).reset_index()
        excel_bytes = export_to_excel(
            sample_orders_df, [cat_agg], ["品类汇总"],
            title="测试报告"
        )
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        sheet_names = wb.sheetnames
        assert '报告信息' in sheet_names
        assert '原始数据' in sheet_names
        assert '品类汇总' in sheet_names

    def test_export_empty_raw_df(self, empty_df):
        excel_bytes = export_to_excel(empty_df, [], [], title="空报告")
        assert isinstance(excel_bytes, bytes)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        assert '报告信息' in wb.sheetnames

    def test_export_long_sheet_name(self, sample_orders_df):
        cat_agg = sample_orders_df.groupby('product_category').agg(
            销售额=('amount', 'sum')
        ).reset_index()
        long_name = "这是一个非常非常长的Sheet名称超过了三十一个字符的限制"
        excel_bytes = export_to_excel(
            sample_orders_df, [cat_agg], [long_name],
            title="长名称测试"
        )
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        assert any('长' in s for s in wb.sheetnames)

    def test_export_report_info_sheet(self, sample_orders_df):
        excel_bytes = export_to_excel(
            sample_orders_df, [], [],
            title="测试标题"
        )
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        info_sheet = wb['报告信息']
        assert info_sheet['A1'].value == "测试标题"
        assert '生成时间' in info_sheet['A2'].value
        assert '数据记录数' in info_sheet['A3'].value


class TestHtmlReport:
    def test_generate_html_basic(self):
        title, subtitle = "测试报告", "时间: 2025-01-01"
        html = generate_html_report(
            "概览", title, subtitle, {}, {}, {}
        )
        assert "测试报告" in html
        assert "时间: 2025-01-01" in html
        assert "<!DOCTYPE html>" in html
        assert "@media print" in html

    def test_generate_html_with_kpi(self):
        kpi_data = {"总销售额": "¥100,000", "订单数": "500 单"}
        html = generate_html_report("概览", "报告", "副标题", kpi_data, {}, {})
        assert "¥100,000" in html
        assert "500 单" in html
        assert "kpi-card" in html

    def test_generate_html_with_tables(self):
        df = pd.DataFrame({'品类': ['A', 'B'], '销售额': [100, 200]})
        html = generate_html_report("概览", "报告", "副标题", {}, {}, {"品类汇总": df})
        assert "品类汇总" in html
        assert "data-table" in html

    def test_generate_html_with_empty_table(self):
        html = generate_html_report("概览", "报告", "副标题", {}, {}, {"空表": pd.DataFrame()})
        assert "空表" not in html

    def test_export_html_report_bytes(self):
        html_content = "<html><body>测试</body></html>"
        result = export_html_report(html_content)
        assert isinstance(result, bytes)
        assert b"测试" in result

    def test_generate_html_figure_placeholder(self):
        import plotly.graph_objects as go
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=[1], y=[1]))

        with patch('utils.exporter.fig_to_base64', return_value=None):
            html = generate_html_report(
                "概览", "报告", "副标题", {},
                {"测试图表": fig}, {}
            )
            assert "图表无法自动渲染" in html


class TestIsBlankImage:
    def test_none_is_blank(self):
        assert _is_blank_image(None) is True

    def test_very_small_is_blank(self):
        assert _is_blank_image(b'\x00' * 500) is True

    def test_reasonable_size_not_blank(self):
        assert _is_blank_image(b'\x00' * 10000) is False

    def test_exactly_threshold(self):
        assert _is_blank_image(b'\x00' * 4999) is True
        assert _is_blank_image(b'\x00' * 5000) is False
